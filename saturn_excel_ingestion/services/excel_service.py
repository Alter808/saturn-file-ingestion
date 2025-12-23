import os
import uuid
from openpyxl import load_workbook
from validators.sheet_validator import validate_sheets
from row_parser import parse_rows

TMP_DIR = "/tmp"

def save_excel_to_tmp(file_bytes, original_filename="upload.xlsx"):
    ext = os.path.splitext(original_filename)[1] or ".xlsx"
    tmp_name = f"upload_{uuid.uuid4().hex}{ext}"
    tmp_path = os.path.join(TMP_DIR, tmp_name)

    with open(tmp_path, "wb") as f:
        f.write(file_bytes)

    return tmp_path

def analyze_excel(tmp_path, rules):
    # Validate sheets against rules
    wb = load_workbook(tmp_path, data_only=True)    
    sheet_result = validate_sheets(wb, rules)
    
    # Cleanup
    wb.close()

    # Returns validation result
    return sheet_result


def read_excel_basic_info(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active  # first sheet
    sheet_name = ws.title

    # headers (row 1)
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = [h for h in first_row if h is not None]

    # count rows (approx reliable in read_only)
    # ws.max_row works fine for read_only mode
    row_count = ws.max_row
    col_count = ws.max_column

    wb.close()

    return {
        "sheet_name": sheet_name,
        "row_count": row_count,
        "col_count": col_count,
        "headers": headers
    }

def process_valid_sheets(tmp_path, sheet_result):
    """
    Process valid sheets
    and apply row validation.
    """
    wb = load_workbook(tmp_path, data_only=True)

    results = []

    for sheet_meta in sheet_result.get("valid_sheets", []):
        sheet_rows = parse_rows(wb, sheet_meta)
        results.append(sheet_rows)

    wb.close()

    return results